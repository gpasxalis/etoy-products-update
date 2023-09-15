from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import os
from sys import exit as sexit
from tkinter import *
from tkinter import ttk, filedialog
import customtkinter
from functools import partial
import threading
import numpy
import csv
import pandas as pd
from PIL import Image

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("./etoy-purple.json")  # Themes: "blue" (standard), "green", "dark-blue"


def create_excel(output_sheet):
    output_sheet["A1"] = "Κωδικός Μοναδικός Soft1"
    output_sheet['A1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['A1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["B1"] = "Υπολ."
    output_sheet['B1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['B1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["C1"] = "Τιμή"
    output_sheet['C1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['C1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["D1"] = "Status"
    output_sheet['D1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['D1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["E1"] = "specific shop"
    output_sheet['E1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['E1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["F1"] = "specific country"
    output_sheet['F1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['F1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["G1"] = "specific group"
    output_sheet['G1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['G1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["H1"] = "specific product price"
    output_sheet['H1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['H1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["I1"] = "specific price reduction"
    output_sheet['I1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['I1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["J1"] = "reductiontype"
    output_sheet['J1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['J1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["K1"] = "Αποθήκες"
    output_sheet['K1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['K1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet["L1"] = "Καταστήματα"
    output_sheet['L1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['L1'].alignment = Alignment(horizontal='center', wrap_text=True)

    output_sheet["M1"] = "Ειδικές Τιμές"
    output_sheet['M1'].font = Font(name='Tahoma', size=8, bold=True)
    output_sheet['M1'].alignment = Alignment(horizontal='center', wrap_text=True)


    output_sheet.column_dimensions['A'].width = 26
    output_sheet.column_dimensions['B'].width = 11
    output_sheet.column_dimensions['C'].width = 11
    output_sheet.column_dimensions['D'].width = 11
    output_sheet.column_dimensions['E'].width = 18
    output_sheet.column_dimensions['F'].width = 18
    output_sheet.column_dimensions['G'].width = 18
    output_sheet.column_dimensions['H'].width = 18
    output_sheet.column_dimensions['I'].width = 20
    output_sheet.column_dimensions['J'].width = 16
    output_sheet.column_dimensions['K'].width = 30
    output_sheet.column_dimensions['L'].width = 60
    output_sheet.column_dimensions['M'].width = 30




def output_excel(import_sheet, path, output_sheet, output_workbook):

    
    create_excel(output_sheet)


    importCurrentRow = 8
    outputCurrentRow = 2
    for eachRow in import_sheet.iter_rows():
        soft1_kodikos_B = import_sheet.cell(row=importCurrentRow, column=2).value
        ypol_G = import_sheet.cell(row=importCurrentRow, column=7).value
        palia_timi_I = import_sheet.cell(row=importCurrentRow, column=9).value
        lianikis_H = import_sheet.cell(row=importCurrentRow, column=8).value
        nea_timi_prosforas_J = import_sheet.cell(row=importCurrentRow, column=8).value
        mak_K = import_sheet.cell(row=importCurrentRow, column=11).value
        ben_L = import_sheet.cell(row=importCurrentRow, column=12).value
        kom_M = import_sheet.cell(row=importCurrentRow, column=13).value



        output_sheet.cell(row=outputCurrentRow, column=1).value = soft1_kodikos_B  
        
        cell_A = output_sheet.cell(row=outputCurrentRow, column=1).value

        if cell_A is not None:

            output_sheet.cell(row=outputCurrentRow, column=2).value = ypol_G

            if (palia_timi_I is not None and palia_timi_I > 0):
                output_sheet.cell(row=outputCurrentRow, column=3).value = palia_timi_I
            else:
                output_sheet.cell(row=outputCurrentRow, column=3).value = lianikis_H

            if ypol_G is not None and ypol_G > 0:
                status = 1
                output_sheet.cell(row=outputCurrentRow, column=4).value = status
            else:
                status = 0
                output_sheet.cell(row=outputCurrentRow, column=4).value = status
            
            
            if (palia_timi_I is not None and palia_timi_I > lianikis_H):
                output_sheet.cell(row=outputCurrentRow, column=13).value = "sales"
            else:
                output_sheet.cell(row=outputCurrentRow, column=13).value = 0


            if palia_timi_I is not None and nea_timi_prosforas_J is not None and palia_timi_I-nea_timi_prosforas_J > 0:
                specific_shop = 0
                specific_country = 0
                specific_group = 0
                specific_product_price = -1
                reductiontype = "amount"
                output_sheet.cell(row=outputCurrentRow, column=5).value = specific_shop
                output_sheet.cell(row=outputCurrentRow, column=6).value = specific_country
                output_sheet.cell(row=outputCurrentRow, column=7).value = specific_group
                output_sheet.cell(row=outputCurrentRow, column=8).value = specific_product_price
                output_sheet.cell(row=outputCurrentRow, column=10).value = reductiontype


                if palia_timi_I is not None and palia_timi_I > 0:
                    specific_price_reduction = palia_timi_I - nea_timi_prosforas_J
                    output_sheet.cell(row=outputCurrentRow, column=9).value = specific_price_reduction
                else:
                    specific_price_reduction = lianikis_H - nea_timi_prosforas_J
                    output_sheet.cell(row=outputCurrentRow, column=9).value = specific_price_reduction

            else:
                specific_shop = ""
                specific_country = ""
                specific_group = ""
                specific_product_price = ""
                reductiontype = ""
                output_sheet.cell(row=outputCurrentRow, column=5).value = specific_shop
                output_sheet.cell(row=outputCurrentRow, column=6).value = specific_country
                output_sheet.cell(row=outputCurrentRow, column=7).value = specific_group
                output_sheet.cell(row=outputCurrentRow, column=8).value = specific_product_price
                output_sheet.cell(row=outputCurrentRow, column=10).value = reductiontype

            
            if mak_K is not None and kom_M is None and ben_L is None:
                apothikes = "MAK:"+str(mak_K)+" BEN:"+str(0)+" KOM:"+str(0)

            elif ben_L is not  None and mak_K is None and kom_M is None:
                apothikes = "MAK:"+str(0)+" BEN:"+str(ben_L)+" KOM:"+str(0)

            elif kom_M is not None and mak_K is None and ben_L is None:
                apothikes = "MAK:"+str(0)+" BEN:"+str(0)+" KOM:"+str(kom_M)

            elif mak_K is not None and ben_L is not None and kom_M is None:
                apothikes = "MAK:"+str(mak_K)+" BEN:"+str(ben_L)+" KOM:"+str(0)
    
            elif mak_K is not None and kom_M is not None and ben_L is None:
                apothikes = "MAK:"+str(mak_K)+" BEN:"+str(0)+" KOM:"+str(kom_M)

            elif ben_L is not None and kom_M is not None and mak_K is None:
                apothikes = "MAK:"+str(0)+" BEN:"+str(ben_L)+" KOM:"+str(kom_M)
        
            elif mak_K is not None and ben_L is not None and kom_M is not None:
                apothikes = "MAK:"+str(mak_K)+" BEN:"+str(ben_L)+" KOM:"+str(kom_M)
    
            elif kom_M is None and mak_K is None and ben_L is None:
                apothikes = "MAK:"+str(0)+" BEN:"+str(0)+" KOM:"+str(0)


            output_sheet.cell(row=outputCurrentRow, column=11).value = apothikes



            if (mak_K is None or mak_K < 0) and (ben_L is None or ben_L < 0) and kom_M is not None and kom_M > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Kosmopolis Κομοτηνή"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            
            elif (mak_K is None or mak_K < 0) and ben_L is not None and (kom_M is None or kom_M < 0) and ben_L > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Βενιζέλου 44 Αλεξανδρούπολη<br>"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            
            elif mak_K is not None and (ben_L is None or ben_L < 0) and (kom_M is None or kom_M < 0) and mak_K > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Λ.Μάκρης 23 Αλεξανδρούπολη<br>"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata

            elif (mak_K is None or mak_K < 0) and ben_L is not None and kom_M is not None and ben_L > 0 and kom_M > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Βενιζέλου 44 Αλεξανδρούπολη<br>Kosmopolis Κομοτηνή"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            
            elif mak_K is not None and (ben_L is None or ben_L < 0) and kom_M is not None and mak_K > 0 and kom_M > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Λ.Μάκρης 23 Αλεξανδρούπολη<br>Kosmopolis Κομοτηνή"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            
            elif mak_K is not None and ben_L is not None and (kom_M is None or kom_M < 0) and mak_K > 0 and ben_L > 0:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Λ.Μάκρης 23 Αλεξανδρούπολη<br>Βενιζέλου 44 Αλεξανδρούπολη<br>"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            
            elif (mak_K is not None and ben_L is not None and kom_M is not None) and (mak_K > 0 and ben_L > 0 and kom_M > 0):
                katastimata = "Διαθέσιμο στα καταστήματα: <br>Λ.Μάκρης 23 Αλεξανδρούπολη<br>Βενιζέλου 44 Αλεξανδρούπολη<br>Kosmopolis Κομοτηνή"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata
            else:
                katastimata = "Διαθέσιμο στα καταστήματα: <br>"
                output_sheet.cell(row=outputCurrentRow, column=12).value = katastimata


        importCurrentRow += 1
        outputCurrentRow +=1


    output_workbook.save(path)
    



class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Product Import | Paschalis Grammenos")
        self.geometry(f"{1280}x{720}")
        self.iconbitmap(r"excel_ico_black.ico")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        self.label_file_explorer = customtkinter.CTkLabel(self, 
                                                          text="Enter name for output file and browse for the input", 
                                                          font=customtkinter.CTkFont(size=20, weight="bold"))   
        
        self.label_file_explorer.grid(row=0, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")

        # create sidebar frame with widgets

        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)))
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "logo_etoy.png")), size=(150, 79))
        

        self.navigation_frame = customtkinter.CTkFrame(self, width=200, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, 
                                                             text=" ", 
                                                             image=self.logo_image,
                                                             compound="left", 
                                                             font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)
        
        self.button_explore = customtkinter.CTkButton(self.navigation_frame, 
                                                      text="Browse" , 
                                                      command = self.explore_files)
        

        self.button_explore.grid(row=1, column=0, padx=20, pady=10)
        self.appearance_mode_label = customtkinter.CTkLabel(self.navigation_frame, 
                                                            text="Appearance Mode:", 
                                                            anchor="w")
        
        self.appearance_mode_label.grid(row=7, column=0, padx=20, pady=(10, 10))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))


        self.output_entry = customtkinter.CTkEntry(self, 
                                                   placeholder_text="Enter output file name")
        self.output_entry.grid(row=3, column=1, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.exit_button = customtkinter.CTkButton(master=self,text="Exit" , 
                                                   fg_color="#e30909",
                                                   command = self.exit, 
                                                   hover_color="#8a0a0a", 
                                                   text_color=("#f5f5f7", "#f5f5f7"))
        
        self.exit_button.grid(row=3, column=3, padx=(20, 20), pady=(20, 20), sticky="nsew")

        # create textbox
        self.textbox = customtkinter.CTkTextbox(self, width=250)

    # FUNCTION FOR CLOSING WINDOW OR DESTROY IT
    def exit(self):
        app.destroy()
        sexit()

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def sidebar_button_event(self):
        print("sidebar_button click")
    
    def explore_files(self):
        explore = threading.Thread(target=partial(self.browseFiles, self.label_file_explorer))
        explore.start()
        
    
    def run_script(self, filename, label_file_explorer):
        """
        Function to start running the code with a new thread so the window don't freeze
        """
        
        book = load_workbook(filename)
        import_sheet = book.active


        output_workbook = Workbook()
        output_sheet = output_workbook.active

        label_file_explorer.configure(text="⚙️ Running", text_color="#543f84")

        output_name = self.output_entry.get()

        for merge in list(import_sheet.merged_cells):
            import_sheet.unmerge_cells(range_string=str(merge))


        username = os.getlogin()


        # SAVE WORKBOOK TO XLSX
        if not output_name:
            path = f'C:\\Users\\{username}\\Desktop\etoy_enimerosi_proionton.xlsx'
        else:
            path = f'C:\\Users\\{username}\\Desktop\\' + output_name + '.xlsx'
        
        output_excel(import_sheet, path, output_sheet, output_workbook)



        # CONVERT XLSX OUTPUT FILE TO CSV FOR INPUT
        if not output_name:
            excel_file = f'C:\\Users\\{username}\\Desktop\etoy_enimerosi_proionton.xlsx'
            read_file = pd.read_excel(excel_file)
            read_file.to_csv (f'C:\\Users\\{username}\\Desktop\etoy_enimerosi_proionton_CSV.csv', index = None, header=True, encoding='utf-8-sig', sep = ';', decimal= ",", quoting=csv.QUOTE_ALL, float_format=None)
        else:
            excel_file = f'C:\\Users\\{username}\\Desktop\\{output_name}.xlsx'
            read_file = pd.read_excel(excel_file)
            read_file.to_csv (f'C:\\Users\\{username}\\Desktop\\{output_name}_CSV.csv', index = None, header=True, encoding='utf-8-sig', sep = ';', decimal= ",", float_format=None)


        # DELETE EXCEL FILE THAT IT'S CREATED WITH OPENPYXL   
        #os.remove(excel_file)




        label_file_explorer.configure(text="✔️ Done", text_color= "green3")
        label_file_explorer.after(2000, self.closing_window(label_file_explorer))
        self.after(2000,lambda:self.destroy())


    def browseFiles(self, label_file_explorer):
        """
        Browse for input excel file in xlsx format
        """

        filename = filedialog.askopenfilename(initialdir = "/",
                                            title = "Select a File",
                                            filetypes = (("Excel file",
                                                            "*.xlsx*"),
                                                        ("all files",
                                                            "*.*")))

        if filename:
            try:
                filename = r"{}".format(filename)
                df = pd.read_excel(filename)
                # Create a Treeview widget
                
                columns = ("Α/Α", "Κωδικός Μοναδικός Soft1", "id shop παλιο", "Online Id Shop", "Εμφάνιση στο eShop", "Κωδ.εργοστασίου", "Υπολ.", "Λιανικής", "παλιά Τιμή", "Νέα τιμή προσφορας", "ΜΑΚ", "ΒΕΝ", 
                                   "ΚΟΜ", "Λιανική χωρίς ΦΠΑ, τη συμπληρωνουμε", "Φ.Π.Α. το σβηνουμε", "Περιγραφή να τη σβηνουμε", "Web page", "AX1", "Διάσταση 2")
                tree = ttk.Treeview(self, columns=columns, show='headings', height="10")
                tree.column("#1")
                tree.column("#2")
                tree.column("#3")
                tree.column("#4")
                tree.column("#5")
                tree.column("#6")
                tree.column("#7")
                tree.column("#8")
                tree.column("#9")
                tree.column("#10")
                tree.column("#11")
                tree.column("#12")
                tree.column("#13")
                tree.column("#14")
                tree.column("#15")
                tree.column("#16")
                tree.column("#17")
                tree.column("#18")
                tree.column("#19")
                tree.heading("#1", text="Α/Α")
                tree.heading("#2", text="Κωδικός Μοναδικός Soft1")
                tree.heading("#3", text="id shop παλιο")
                tree.heading("#4", text="Online Id Shop")
                tree.heading("#5", text="Εμφάνιση στο eShop")
                tree.heading("#6", text="Κωδ.εργοστασίου")
                tree.heading("#7", text="Υπολ.")
                tree.heading("#8", text="Λιανικής")
                tree.heading("#9", text="παλιά Τιμή")
                tree.heading("#10", text="Νέα τιμή προσφορας")
                tree.heading("#11", text="ΜΑΚ")
                tree.heading("#12", text="ΒΕΝ")
                tree.heading("#13", text="ΚΟΜ")
                tree.heading("#14", text="Λιανική χωρίς ΦΠΑ, τη συμπληρωνουμε")
                tree.heading("#15", text="Φ.Π.Α. το σβηνουμε")
                tree.heading("#16", text="Περιγραφή να τη σβηνουμε")
                tree.heading("#17", text="Web page")
                tree.heading("#18", text="AX1")
                tree.heading("#19", text="Διάσταση 2")


                treeXScroll = ttk.Scrollbar(self,orient=HORIZONTAL)
                treeXScroll.configure(command=tree.xview)
                tree.configure(xscrollcommand=treeXScroll.set)

                # Clear all the previous data in tree
                self.clear_treeview(tree)

                # Put Data in Rows
                df_rows = df.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                tree.grid(row=1, column=1, padx=(20, 0), pady=(2, 0), sticky="nsew")
                
                treeXScroll.grid(row=2, column=1, padx=(20, 0), pady=(1, 0), sticky="nsew")

            except ValueError:
                label_file_explorer.config(text="❌ File could not be opened", 
                                           text_color='red2')
            except FileNotFoundError:
                label_file_explorer.config(text="❌ File Not Found", 
                                           text_color='red2')
        
        
        if filename is None or filename=='':
            label_file_explorer.configure(text="⚠️ Browse window closed", 
                                          text_color='#ad8d0e')
        else:
            label_file_explorer.configure(text="✔️ File Found", 
                                          text_color="green3")
    
            self.button_run = customtkinter.CTkButton(self.navigation_frame,
                                                      text = "Run Script", 
                                                      command = threading.Thread(target=partial(self.run_script, filename, label_file_explorer)).start,
                                                      fg_color="#23ad31",
                                                      text_color="#f5f5f7",
                                                      hover_color="#167020")
            self.button_run.grid(row=2, column=0, padx=20, pady=10)
            

            
    
    def closing_window(self, label_file_explorer):
        """
        Change Label for widget in tkinter field
        """

        label_file_explorer.configure(text="❌ Closing the window", 
                                      text_color='red2')
    
    # Clear the Treeview Widget
    def clear_treeview(self, tree):
        tree.delete(*tree.get_children())




if __name__ == "__main__":
    app = App()
    app.mainloop()